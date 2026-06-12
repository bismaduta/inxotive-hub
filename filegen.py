"""
File Generator — Excel (.xlsx) and PDF file generation for INXOTIVE HUB.
Zero heavy external dependencies; falls back gracefully with install instructions.

Functions:
    generate_excel(data)   — Create Excel workbook from headers+rows
    generate_pdf(data)     — Create simple PDF from plain text content
    generate_invoice(data) — Create professional invoice PDF with items table
    cleanup_old_files()    — Remove generated files older than 1 hour
"""

import asyncio
import logging
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger("filegen")

# ── Output directory ──
OUTPUT_DIR = Path("/tmp/inxotive-files")

# ── Optional dependency: openpyxl ──
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Optional dependency: reportlab ──
try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


# ── Helpers ──

def _get_output_path(ext: str) -> str:
    """Generate a unique file path under /tmp/inxotive-files/.

    Args:
        ext: File extension including the dot, e.g. '.xlsx' or '.pdf'.

    Returns:
        Absolute path string.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stem = uuid.uuid4().hex[:16]
    return str(OUTPUT_DIR / f"{stem}{ext}")


def _now() -> datetime:
    """Return current UTC datetime (timezone-aware)."""
    return datetime.now(timezone.utc)


def _filename_from_timestamp(prefix: str, ext: str) -> str:
    """Build a human-readable filename with a timestamp."""
    ts = _now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}{ext}"


# ── Cleanup ──

async def cleanup_old_files() -> Dict[str, Any]:
    """Remove generated files older than 1 hour from the output directory.

    Runs asynchronously — the I/O is fast enough on a local tmpfs that a
    simple thread-pool executor is not warranted, but the function signature
    is async so it can be awaited by the callers.

    Returns:
        {"success": bool, "cleaned": int, "error": str}
    """
    result: Dict[str, Any] = {"success": True, "cleaned": 0, "error": ""}
    cutoff = _now() - timedelta(hours=1)

    try:
        if not OUTPUT_DIR.exists():
            return result

        removed = 0
        for fpath in OUTPUT_DIR.iterdir():
            if not fpath.is_file():
                continue
            try:
                mtime = datetime.fromtimestamp(fpath.stat().st_mtime, tz=timezone.utc)
                if mtime < cutoff:
                    fpath.unlink()
                    removed += 1
                    logger.debug("Cleaned old file: %s", fpath.name)
            except OSError as exc:
                logger.warning("Could not remove %s: %s", fpath.name, exc)

        result["cleaned"] = removed
        if removed:
            logger.info("Cleanup removed %d expired file(s)", removed)
    except Exception as exc:
        logger.error("Cleanup error: %s", exc)
        result["success"] = False
        result["error"] = str(exc)

    return result


# ── Excel Generation ──

async def generate_excel(data: Dict[str, Any]) -> Dict[str, Any]:
    """Create an Excel (.xlsx) workbook from structured data.

    Expected data format::

        {
            "headers": ["Name", "Age", "City"],
            "rows": [
                ["Alice", 30, "Jakarta"],
                ["Bob", 25, "Bandung"],
            ],
            "sheet_name": "Sheet1",        # optional, defaults to "Sheet1"
        }

    Returns:
        {"success": bool, "path": str, "error": str}
    """
    result: Dict[str, Any] = {"success": False, "path": "", "error": ""}

    if not OPENPYXL_AVAILABLE:
        msg = (
            "openpyxl is not installed. Install it with: "
            "pip install openpyxl"
        )
        logger.error(msg)
        result["error"] = msg
        return result

    headers: List[str] = data.get("headers", [])
    rows: List[List[Any]] = data.get("rows", [])
    sheet_name: str = data.get("sheet_name", "Sheet1")

    if not headers and not rows:
        msg = "No data provided: 'headers' and 'rows' are both empty"
        logger.error(msg)
        result["error"] = msg
        return result

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name max 31 chars

        # ── Styles ──
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            bottom=Side(style="thin", color="475569"),
        )

        cell_font = Font(name="Calibri", size=11)
        cell_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # ── Write headers ──
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border

        # ── Write rows ──
        start_row = 2 if headers else 1
        for row_idx, row_data in enumerate(rows, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border

        # ── Auto-fit column widths ──
        for col_idx in range(1, max(len(headers) if headers else 1, max((len(r) for r in rows), default=1)) + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            # Check header length
            if headers and col_idx <= len(headers):
                max_length = len(str(headers[col_idx - 1]))

            # Check row values
            for row_data in rows:
                if col_idx <= len(row_data):
                    val_len = len(str(row_data[col_idx - 1]))
                    if val_len > max_length:
                        max_length = val_len

            # Cap at 60 chars and add padding
            adjusted = min(max_length + 3, 60)
            ws.column_dimensions[col_letter].width = max(adjusted, 8)

        # ── Freeze top row if headers exist ──
        if headers:
            ws.freeze_panes = "A2"

        # ── Output path ──
        output_path = _get_output_path(".xlsx")
        wb.save(output_path)
        logger.info("Excel generated: %s", output_path)

        result["success"] = True
        result["path"] = output_path

    except Exception as exc:
        logger.error("Excel generation error: %s", exc, exc_info=True)
        result["error"] = f"Excel generation failed: {exc}"

    return result


# ── PDF Generation ──

async def generate_pdf(data: Dict[str, Any]) -> Dict[str, Any]:
    """Create a simple PDF document from text content.

    Expected data format::

        {
            "text": "Full document text here...\\n\\nSecond paragraph.",
            "title": "Document Title",     # optional
            "filename": "report",           # optional stem (without extension)
        }

    Returns:
        {"success": bool, "path": str, "error": str}
    """
    result: Dict[str, Any] = {"success": False, "path": "", "error": ""}

    if not REPORTLAB_AVAILABLE:
        msg = (
            "reportlab is not installed. Install it with: "
            "pip install reportlab"
        )
        logger.error(msg)
        result["error"] = msg
        return result

    text: str = data.get("text", "")
    title: str = data.get("title", "")

    if not text.strip() and not title.strip():
        msg = "No content provided: 'text' and 'title' are both empty"
        logger.error(msg)
        result["error"] = msg
        return result

    try:
        filename = data.get("filename", "document")
        output_path = str(OUTPUT_DIR / _filename_from_timestamp(filename, ".pdf"))
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
            leftMargin=2.5 * cm,
            rightMargin=2.5 * cm,
        )

        styles = getSampleStyleSheet()
        story = []

        # ── Title ──
        if title:
            title_style = ParagraphStyle(
                "DocTitle",
                parent=styles["Title"],
                fontSize=18,
                spaceAfter=12,
                textColor=colors.HexColor("#1E293B"),
            )
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 6 * mm))

        # ── Content paragraphs ──
        body_style = ParagraphStyle(
            "DocBody",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            spaceAfter=8,
            textColor=colors.HexColor("#334155"),
        )

        paragraphs = text.split("\n\n")
        for para in paragraphs:
            stripped = para.strip()
            if stripped:
                # Escape XML special chars for reportlab
                safe = stripped.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                # Simple markdown-like bold/italic conversion
                safe = safe.replace("&lt;b&gt;", "<b>").replace("&lt;/b&gt;", "</b>")
                safe = safe.replace("&lt;i&gt;", "<i>").replace("&lt;/i&gt;", "</i>")
                story.append(Paragraph(safe, body_style))

        doc.build(story)
        logger.info("PDF generated: %s", output_path)

        result["success"] = True
        result["path"] = output_path

    except Exception as exc:
        logger.error("PDF generation error: %s", exc, exc_info=True)
        result["error"] = f"PDF generation failed: {exc}"

    return result


# ── Invoice PDF Generation ──

_INVOICE_COMPANY_NAME = "INXOTIVE OFFICE"
_INVOICE_COMPANY_ADDR = "Indonesia"
_INVOICE_FOOTER = "Thank you for your business!"


async def generate_invoice(invoice_data: Dict[str, Any]) -> Dict[str, Any]:
    """Generate a professional invoice PDF with items table.

    Expected ``invoice_data`` format::

        {
            "client_name": "PT Contoh Sejahtera",      # required
            "client_address": "Jl. Merdeka No. 1",    # optional
            "invoice_number": "INV-2024-001",          # optional (auto)
            "date": "2024-01-15",                      # optional (today)
            "items": [                                  # required
                {
                    "description": "Web Development",
                    "qty": 1,
                    "rate": 5000000,
                    "amount": 5000000,
                },
            ],
            "notes": "Payment due within 14 days.",    # optional
            "tax_rate": 11,                            # optional, percentage
        }

    Returns:
        {"success": bool, "path": str, "error": str}
    """
    result: Dict[str, Any] = {"success": False, "path": "", "error": ""}

    if not REPORTLAB_AVAILABLE:
        msg = (
            "reportlab is not installed. Install it with: "
            "pip install reportlab"
        )
        logger.error(msg)
        result["error"] = msg
        return result

    # ── Validate required fields ──
    client_name: str = invoice_data.get("client_name", "").strip()
    if not client_name:
        msg = "Missing required field: 'client_name'"
        logger.error(msg)
        result["error"] = msg
        return result

    items: List[Dict[str, Any]] = invoice_data.get("items", [])
    if not items:
        msg = "Missing required field: 'items' (must be a non-empty list)"
        logger.error(msg)
        result["error"] = msg
        return result

    # ── Optional fields with defaults ──
    client_address: str = invoice_data.get("client_address", "").strip()
    invoice_number: str = invoice_data.get("invoice_number", "").strip()
    if not invoice_number:
        invoice_number = f"INV-{_now().strftime('%Y%m%d-%H%M%S')}"

    date_str: str = invoice_data.get("date", "").strip()
    if not date_str:
        date_str = _now().strftime("%Y-%m-%d")

    notes: str = invoice_data.get("notes", "").strip()
    tax_rate: float = float(invoice_data.get("tax_rate", 0))

    try:
        output_path = str(OUTPUT_DIR / _filename_from_timestamp("invoice", ".pdf"))
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
        )

        story = []
        page_width = A4[0] - 4 * cm  # usable width after margins

        # ── Color palette ──
        DARK_BG = colors.HexColor("#1E293B")
        ACCENT = colors.HexColor("#6366F1")
        LIGHT_BG = colors.HexColor("#F8FAFC")
        MEDIUM_GRAY = colors.HexColor("#64748B")
        DARK_TEXT = colors.HexColor("#1E293B")
        BORDER_COLOR = colors.HexColor("#CBD5E1")

        # ── Styles ──
        styles = getSampleStyleSheet()

        company_style = ParagraphStyle(
            "CompanyName",
            parent=styles["Normal"],
            fontSize=20,
            bold=True,
            textColor=DARK_BG,
            spaceAfter=2,
        )
        company_sub_style = ParagraphStyle(
            "CompanySub",
            parent=styles["Normal"],
            fontSize=9,
            textColor=MEDIUM_GRAY,
            spaceAfter=4,
        )
        label_style = ParagraphStyle(
            "Label",
            parent=styles["Normal"],
            fontSize=8,
            textColor=MEDIUM_GRAY,
            spaceAfter=1,
        )
        value_style = ParagraphStyle(
            "Value",
            parent=styles["Normal"],
            fontSize=10,
            textColor=DARK_TEXT,
            spaceAfter=6,
        )
        header_style = ParagraphStyle(
            "InvoiceHeader",
            parent=styles["Normal"],
            fontSize=24,
            bold=True,
            textColor=ACCENT,
            spaceAfter=4,
        )
        table_header_style = ParagraphStyle(
            "TableHeader",
            parent=styles["Normal"],
            fontSize=9,
            bold=True,
            textColor=colors.white,
            alignment=TA_CENTER,
        )
        table_cell_style = ParagraphStyle(
            "TableCell",
            parent=styles["Normal"],
            fontSize=9,
            textColor=DARK_TEXT,
        )
        table_cell_right = ParagraphStyle(
            "TableCellRight",
            parent=styles["Normal"],
            fontSize=9,
            textColor=DARK_TEXT,
            alignment=TA_RIGHT,
        )
        total_label_style = ParagraphStyle(
            "TotalLabel",
            parent=styles["Normal"],
            fontSize=11,
            bold=True,
            textColor=DARK_TEXT,
        )
        total_value_style = ParagraphStyle(
            "TotalValue",
            parent=styles["Normal"],
            fontSize=11,
            bold=True,
            textColor=DARK_BG,
            alignment=TA_RIGHT,
        )
        notes_style = ParagraphStyle(
            "Notes",
            parent=styles["Normal"],
            fontSize=9,
            textColor=MEDIUM_GRAY,
            leading=13,
        )
        footer_style = ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontSize=8,
            textColor=MEDIUM_GRAY,
            alignment=TA_CENTER,
        )

        # ── Company Header ──
        story.append(Paragraph(_INVOICE_COMPANY_NAME, company_style))
        story.append(Paragraph(_INVOICE_COMPANY_ADDR, company_sub_style))
        story.append(Spacer(1, 4 * mm))

        # ── Divider line ──
        divider_data = [[""]]
        divider_table = Table(divider_data, colWidths=[page_width])
        divider_table.setStyle(TableStyle([
            ("LINEBELOW", (0, 0), (-1, -1), 1, ACCENT),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(divider_table)
        story.append(Spacer(1, 4 * mm))

        # ── Invoice Title & Number (side by side) ──
        # Using a two-column table for INVOICE label and invoice metadata
        half_w = page_width / 2
        inv_header_data = [
            [
                Paragraph("INVOICE", header_style),
                Paragraph(
                    f"<b>{invoice_number}</b><br/>"
                    f"<font size='8' color='{MEDIUM_GRAY.hexval()}'>Date: {date_str}</font>",
                    ParagraphStyle(
                        "InvMeta",
                        parent=styles["Normal"],
                        fontSize=10,
                        textColor=DARK_TEXT,
                        alignment=TA_RIGHT,
                    ),
                ),
            ]
        ]
        inv_header_table = Table(inv_header_data, colWidths=[half_w, half_w])
        inv_header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(inv_header_table)
        story.append(Spacer(1, 8 * mm))

        # ── Bill To ──
        bill_to_lines = [f"<b>{client_name}</b>"]
        if client_address:
            bill_to_lines.append(client_address)
        bill_to_html = "<br/>".join(bill_to_lines)

        bill_to_data = [
            [Paragraph("BILL TO", label_style)],
            [Paragraph(bill_to_html, value_style)],
        ]
        bill_to_table = Table(bill_to_data, colWidths=[page_width])
        bill_to_table.setStyle(TableStyle([
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("BACKGROUND", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 1), (-1, 1), LIGHT_BG),
        ]))
        story.append(bill_to_table)
        story.append(Spacer(1, 6 * mm))

        # ── Items Table ──
        col_widths = [page_width * 0.40, page_width * 0.12, page_width * 0.22, page_width * 0.26]
        table_headers = [
            Paragraph("DESCRIPTION", table_header_style),
            Paragraph("QTY", table_header_style),
            Paragraph("RATE", table_header_style),
            Paragraph("AMOUNT", table_header_style),
        ]

        table_data = [table_headers]

        for item in items:
            desc = item.get("description", "")
            qty = str(item.get("qty", 1))
            rate = _format_currency(item.get("rate", 0))
            amount = _format_currency(item.get("amount", 0))
            table_data.append([
                Paragraph(desc, table_cell_style),
                Paragraph(qty, ParagraphStyle("Qty", parent=table_cell_style, alignment=TA_CENTER)),
                Paragraph(rate, table_cell_right),
                Paragraph(amount, table_cell_right),
            ])

        # ── Totals ──
        subtotal = sum(item.get("amount", item.get("qty", 1) * item.get("rate", 0)) for item in items)
        tax_amount = subtotal * (tax_rate / 100) if tax_rate > 0 else 0
        grand_total = subtotal + tax_amount

        # Add blank row before totals
        table_data.append([
            Paragraph("", table_cell_style),
            Paragraph("", table_cell_style),
            Paragraph("", table_cell_style),
            Paragraph("", table_cell_style),
        ])

        # Subtotal row
        table_data.append([
            Paragraph("", table_cell_style),
            Paragraph("", table_cell_style),
            Paragraph("<b>Subtotal</b>", total_label_style),
            Paragraph(f"<b>{_format_currency(subtotal)}</b>", total_value_style),
        ])

        # Tax row
        if tax_rate > 0:
            table_data.append([
                Paragraph("", table_cell_style),
                Paragraph("", table_cell_style),
                Paragraph(f"<b>Tax ({tax_rate}%)</b>", total_label_style),
                Paragraph(f"<b>{_format_currency(tax_amount)}</b>", total_value_style),
            ])

        # Grand Total row
        table_data.append([
            Paragraph("", table_cell_style),
            Paragraph("", table_cell_style),
            Paragraph("<b>TOTAL</b>", ParagraphStyle("TotalBold", parent=total_label_style, fontSize=13)),
            Paragraph(
                f"<b>{_format_currency(grand_total)}</b>",
                ParagraphStyle("TotalValueBold", parent=total_value_style, fontSize=13, textColor=ACCENT),
            ),
        ])

        items_table = Table(table_data, colWidths=col_widths)
        items_table.setStyle(TableStyle([
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0), DARK_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            # Grid
            ("GRID", (0, 0), (-1, -4), 0.5, BORDER_COLOR),  # items grid
            ("LINEBELOW", (0, -1), (-1, -1), 2, ACCENT),    # grand total underline
            # Alignment
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            # Alternating row colors for items (skip header)
            ("BACKGROUND", (0, 1), (-1, 1), LIGHT_BG),
            ("BACKGROUND", (0, 2), (-1, 2), colors.white),
            ("BACKGROUND", (0, 3), (-1, 3), LIGHT_BG),
            # Summary rows (white background)
            ("BACKGROUND", (0, -4), (-1, -1), colors.white),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 8 * mm))

        # ── Notes ──
        if notes:
            notes_data = [
                [Paragraph("NOTES", label_style)],
                [Paragraph(notes, notes_style)],
            ]
            notes_table = Table(notes_data, colWidths=[page_width])
            notes_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), LIGHT_BG),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(notes_table)
            story.append(Spacer(1, 6 * mm))

        # ── Footer ──
        story.append(Spacer(1, 6 * mm))
        footer_data = [[""]]
        footer_divider = Table(footer_data, colWidths=[page_width])
        footer_divider.setStyle(TableStyle([
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, BORDER_COLOR),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(footer_divider)
        story.append(Paragraph(_INVOICE_FOOTER, footer_style))

        doc.build(story)
        logger.info("Invoice PDF generated: %s", output_path)

        result["success"] = True
        result["path"] = output_path

    except Exception as exc:
        logger.error("Invoice generation error: %s", exc, exc_info=True)
        result["error"] = f"Invoice generation failed: {exc}"

    return result


# ── Utility ──

def _format_currency(amount: float) -> str:
    """Format a number as IDR currency string."""
    if amount >= 1_000_000_000:
        return f"Rp{amount / 1_000_000_000:,.2f}B"
    elif amount >= 1_000_000:
        return f"Rp{amount / 1_000_000:,.2f}M"
    else:
        return f"Rp{amount:,.0f}"
